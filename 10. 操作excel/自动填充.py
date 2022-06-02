import datetime

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

#填充数字之和
for i in range(1,50):
    for j in range(1,50):
        ws.cell(i,j).value=i+j

#填充文本之和
for i in range(51,100):
    for j in range(1,50):
        ws.cell(i, j).value = str(i) + str(j)


wb.save("sample2.xlsx")