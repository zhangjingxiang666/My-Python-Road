from openpyxl import Workbook
from openpyxl import load_workbook
import os
if __name__ == "__main__":
    rootdir = 'D:/杂/计算机技巧相关/教程/Python爬虫/untitled1/操作excel/test'
    list = os.listdir(rootdir)
    print(list)# 列出文件夹下所有的目录与文件
    wb = Workbook()#创建工作簿
    ws = wb.create_sheet()#在工作簿里创建工作表
    for i in range(0, len(list)):
        path = os.path.join(rootdir, list[i])
        #print(path)
        wb2 = load_workbook(path)
        ws2=wb2['Sheet1']
        #print(ws2['A1'].value)
        #print(len(ws2['B']))
        for j in range (0,len(ws2['B'])):
            m=ws2.cell(row=j+1, column=2)#确定行列时直接用row（行）=....，column（列）=......
            print(m.value)
            ws.cell(row=j+1,column=i+2).value=m.value

wb.save("output.xlsx")
        #colC = ws2['C']

        # # 你想对文件的操作

